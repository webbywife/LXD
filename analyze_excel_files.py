#!/usr/bin/env python3
import openpyxl
import os
from pathlib import Path

# List of Excel files to analyze
excel_files = [
    "EPP_TLE_MATATAG_AI_Reference_Curriculum.xlsx",
    "Makabansa_G1-3_AI_Curriculum_Reference.xlsx",
    "MATATAG_English_CG_AI_Reference.xlsx",
    "MATATAG_Filipino_CG_AI_Reference.xlsx",
    "MATATAG_GMRC_VE_AI_Reference.xlsx",
    "MATATAG_Kindergarten_CG_AI_Reference.xlsx",
    "MATATAG_Language_G1_AI_Reference.xlsx",
    "MATATAG_Math_Curriculum_AI_Reference.xlsx",
    "MATATAG_Music_Arts_AI_Reference.xlsx",
    "MATATAG_PE_Health_AI_Curriculum_Reference.xlsx",
    "MATATAG_RL_G1_AI_Reference.xlsx",
    "MATATAG_Science_CG_AI_Reference.xlsx"
]

base_dir = "/Users/leaabarentos/LXD"

def analyze_excel_file(file_path):
    """Analyze an Excel file and return its structure."""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        file_name = os.path.basename(file_path)

        print(f"\n{'='*80}")
        print(f"FILE: {file_name}")
        print(f"{'='*80}")

        # Get all sheet names
        sheet_names = workbook.sheetnames
        print(f"\nSheet Names: {sheet_names}")
        print(f"Total Sheets: {len(sheet_names)}")

        # Analyze each sheet
        for sheet_name in sheet_names:
            print(f"\n{'-'*80}")
            print(f"SHEET: {sheet_name}")
            print(f"{'-'*80}")

            sheet = workbook[sheet_name]

            # Get total rows and columns
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"Total Rows: {max_row}")
            print(f"Total Columns: {max_col}")

            if max_row == 0:
                print("(Empty sheet)")
                continue

            # Get column headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(cell_value)

            print(f"\nColumn Headers:")
            for i, header in enumerate(headers, 1):
                print(f"  {i}. {header}")

            # Get sample rows (rows 2-4 if they exist)
            print(f"\nSample Data Rows:")
            sample_rows = min(4, max_row)  # Get up to 3 data rows (rows 2-4)

            for row_num in range(2, sample_rows + 1):
                print(f"\n  Row {row_num}:")
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    # Truncate long values for readability
                    if cell_value and isinstance(cell_value, str) and len(cell_value) > 100:
                        cell_value = cell_value[:100] + "..."
                    row_data.append(cell_value)

                for header, value in zip(headers, row_data):
                    if value is not None and value != "":
                        print(f"    {header}: {value}")

        workbook.close()

    except FileNotFoundError:
        print(f"\n{'='*80}")
        print(f"FILE: {os.path.basename(file_path)}")
        print(f"{'='*80}")
        print(f"ERROR: File not found")
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"FILE: {os.path.basename(file_path)}")
        print(f"{'='*80}")
        print(f"ERROR: {str(e)}")

# Main execution
print("ANALYZING EXCEL FILES IN /Users/leaabarentos/LXD/")
print("="*80)

for excel_file in excel_files:
    file_path = os.path.join(base_dir, excel_file)
    analyze_excel_file(file_path)

print(f"\n{'='*80}")
print("ANALYSIS COMPLETE")
print(f"{'='*80}")
