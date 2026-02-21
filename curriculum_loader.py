"""
MATATAG Curriculum Data Loader
Parses all 13 MATATAG Excel spreadsheets into a SQLite database.
"""

import os
import re
import sqlite3
import json
from collections import defaultdict
from openpyxl import load_workbook

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(DATA_DIR, "curriculum.db")

# Map of Excel files to their subject identifiers
SUBJECT_FILES = {
    "Mathematics": "MATATAG_Math_Curriculum_AI_Reference.xlsx",
    "Science": "MATATAG_Science_CG_AI_Reference.xlsx",
    "English": "MATATAG_English_CG_AI_Reference.xlsx",
    "Filipino": "MATATAG_Filipino_CG_AI_Reference.xlsx",
    "GMRC_VE": "MATATAG_GMRC_VE_AI_Reference.xlsx",
    "Kindergarten": "MATATAG_Kindergarten_CG_AI_Reference.xlsx",
    "Language_G1": "MATATAG_Language_G1_AI_Reference.xlsx",
    "Music_Arts": "MATATAG_Music_Arts_AI_Reference.xlsx",
    "PE_Health": "MATATAG_PE_Health_AI_Curriculum_Reference.xlsx",
    "Reading_Literacy_G1": "MATATAG_RL_G1_AI_Reference.xlsx",
    "EPP_TLE": "EPP_TLE_MATATAG_AI_Reference_Curriculum.xlsx",
    "Makabansa": "Makabansa_G1-3_AI_Curriculum_Reference.xlsx",
    "Araling_Panlipunan": "MATATAG_AP_Curriculum_AI_Reference.xlsx",
}

# Display names for subjects
SUBJECT_DISPLAY = {
    "Mathematics": "Mathematics",
    "Science": "Science",
    "English": "English",
    "Filipino": "Filipino",
    "GMRC_VE": "GMRC / Values Education",
    "Kindergarten": "Kindergarten (All Areas)",
    "Language_G1": "Language (Mother Tongue) - Grade 1",
    "Music_Arts": "Music and Arts (MAPEH)",
    "PE_Health": "PE and Health (MAPEH)",
    "Reading_Literacy_G1": "Reading & Literacy - Grade 1",
    "EPP_TLE": "EPP / TLE (Technology & Livelihood)",
    "Makabansa": "Makabansa (Civics/History/Geography)",
    "Araling_Panlipunan": "Araling Panlipunan (Social Studies)",
}


SHS_FILE = "SSHS_Core_Curriculum_AI_Reference.xlsx"

# Maps Subject_Code in the Excel to a unique subject_id in the DB
SHS_SUBJECT_CODES = {
    "EC":  "SHS_Effective_Communication",
    "MK":  "SHS_Mabisang_Komunikasyon",
    "GM":  "SHS_General_Mathematics",
    "GS":  "SHS_General_Science",
    "LCS": "SHS_Life_and_Career_Skills",
    "KLP": "SHS_Kasaysayan_at_Lipunan",
}

SHS_DISPLAY = {
    "SHS_Effective_Communication": "SHS \u2013 Effective Communication (English)",
    "SHS_Mabisang_Komunikasyon":   "SHS \u2013 Mabisang Komunikasyon (Filipino)",
    "SHS_General_Mathematics":     "SHS \u2013 General Mathematics",
    "SHS_General_Science":         "SHS \u2013 General Science",
    "SHS_Life_and_Career_Skills":  "SHS \u2013 Life and Career Skills",
    "SHS_Kasaysayan_at_Lipunan":   "SHS \u2013 Kasaysayan at Lipunan ng Pilipinas",
}


def init_database():
    """Create all tables in the SQLite database."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("DROP TABLE IF EXISTS learning_competencies")
    c.execute("DROP TABLE IF EXISTS standards")
    c.execute("DROP TABLE IF EXISTS pedagogical_approaches")
    c.execute("DROP TABLE IF EXISTS twenty_first_century_skills")
    c.execute("DROP TABLE IF EXISTS crosscutting_concepts")
    c.execute("DROP TABLE IF EXISTS domain_sequence")
    c.execute("DROP TABLE IF EXISTS subjects")

    c.execute("""
        CREATE TABLE subjects (
            id TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            filename TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE learning_competencies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            lc_id TEXT,
            grade TEXT,
            quarter TEXT,
            key_stage TEXT,
            domain TEXT,
            subdomain TEXT,
            content_topic TEXT,
            learning_competency TEXT,
            content_standard TEXT,
            performance_standard TEXT,
            blooms_level TEXT,
            competency_type TEXT,
            ai_tags TEXT,
            extra_data TEXT,
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        )
    """)

    c.execute("""
        CREATE TABLE standards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            standard_type TEXT,
            grade TEXT,
            key_stage TEXT,
            description TEXT,
            extra_data TEXT,
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        )
    """)

    c.execute("""
        CREATE TABLE pedagogical_approaches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            approach_name TEXT,
            description TEXT,
            strategies TEXT,
            extra_data TEXT,
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        )
    """)

    c.execute("""
        CREATE TABLE twenty_first_century_skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            skill_name TEXT,
            category TEXT,
            description TEXT,
            extra_data TEXT,
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        )
    """)

    c.execute("""
        CREATE TABLE crosscutting_concepts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            concept_name TEXT,
            description TEXT,
            extra_data TEXT,
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        )
    """)

    c.execute("""
        CREATE TABLE domain_sequence (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            domain_name TEXT,
            grade TEXT,
            sequence_info TEXT,
            extra_data TEXT,
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        )
    """)

    # Indexes for fast querying
    c.execute("CREATE INDEX idx_lc_subject ON learning_competencies(subject_id)")
    c.execute("CREATE INDEX idx_lc_grade ON learning_competencies(grade)")
    c.execute("CREATE INDEX idx_lc_quarter ON learning_competencies(quarter)")
    c.execute("CREATE INDEX idx_lc_subject_grade ON learning_competencies(subject_id, grade)")
    c.execute("CREATE INDEX idx_lc_subject_grade_quarter ON learning_competencies(subject_id, grade, quarter)")

    conn.commit()
    conn.close()


def _cell_value(cell):
    """Get string value from a cell, handling None."""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _find_sheet(wb, keywords):
    """Find a sheet by matching keywords in the sheet name (case-insensitive)."""
    for name in wb.sheetnames:
        lower = name.lower()
        for kw in keywords:
            if kw.lower() in lower:
                return wb[name]
    return None


def _read_sheet_rows(ws, max_empty=5):
    """Read all rows from a worksheet, stopping after consecutive empty rows."""
    rows = []
    empty_count = 0
    for row in ws.iter_rows(min_row=1):
        vals = [_cell_value(c) for c in row]
        if all(v == "" for v in vals):
            empty_count += 1
            if empty_count >= max_empty:
                break
            continue
        empty_count = 0
        rows.append(vals)
    return rows


def _normalize_header(h):
    """Normalize a header string for matching."""
    return h.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")


def _find_column(headers, *keywords):
    """Find a column index by matching keywords in normalized headers."""
    for i, h in enumerate(headers):
        nh = _normalize_header(h)
        for kw in keywords:
            if kw in nh:
                return i
    return None


def load_learning_competencies(conn, subject_id, ws):
    """Parse and load learning competencies from a worksheet."""
    rows = _read_sheet_rows(ws)
    if len(rows) < 2:
        return 0

    headers = rows[0]
    h_norm = [_normalize_header(h) for h in headers]

    # Flexible column mapping
    col_lc_id = _find_column(headers, "lc_id", "lc_code")
    col_grade = _find_column(headers, "grade")
    col_quarter = _find_column(headers, "quarter")
    col_key_stage = _find_column(headers, "key_stage", "keystage")
    col_domain = _find_column(headers, "domain", "component", "learning_area")
    col_subdomain = _find_column(headers, "subdomain", "sub_domain", "sub_component")
    col_topic = _find_column(headers, "content_topic", "topic", "theme", "content_focus", "nilalaman")
    col_lc = _find_column(headers, "learning_competency", "competency_text", "competency_description", "kasanayang", "pampagkatuto")
    col_cs = _find_column(headers, "content_standard", "pangnilalaman")
    col_ps = _find_column(headers, "performance_standard", "pagganap")
    col_bloom = _find_column(headers, "bloom", "blooms")
    col_type = _find_column(headers, "competency_type")
    col_tags = _find_column(headers, "ai_tag", "ai_searchable", "tags")

    # Also check for sub-competency columns (GMRC style)
    col_sub_a = _find_column(headers, "sub_competency_a", "sub_competency a")
    col_sub_b = _find_column(headers, "sub_competency_b", "sub_competency b")
    col_sub_c = _find_column(headers, "sub_competency_c", "sub_competency c")

    c = conn.cursor()
    count = 0
    last_main_data = {}  # Track last non-empty main fields for grouped rows
    for row in rows[1:]:
        if len(row) == 0:
            continue

        def get(idx):
            if idx is not None and idx < len(row):
                return row[idx]
            return ""

        lc_text = get(col_lc)

        # For files with sub-competencies (like GMRC), if main LC is empty,
        # try sub-competencies, then fall back to content standard
        if not lc_text:
            sub_parts = []
            for sc in [col_sub_a, col_sub_b, col_sub_c]:
                sv = get(sc)
                if sv:
                    sub_parts.append(sv)
            if sub_parts:
                lc_text = " | ".join(sub_parts)
            elif get(col_cs):
                # Use content standard as the competency description
                lc_text = get(col_cs)
            else:
                continue

        # Track/inherit data for grouped rows
        current_grade = get(col_grade)
        current_quarter = get(col_quarter)
        if current_grade:
            last_main_data["grade"] = current_grade
        if current_quarter:
            last_main_data["quarter"] = current_quarter
        if get(col_key_stage):
            last_main_data["key_stage"] = get(col_key_stage)
        if get(col_domain):
            last_main_data["domain"] = get(col_domain)
        if get(col_topic):
            last_main_data["topic"] = get(col_topic)
        if get(col_cs):
            last_main_data["cs"] = get(col_cs)
        if get(col_ps):
            last_main_data["ps"] = get(col_ps)

        # Build extra data from unmapped columns
        extra = {}
        mapped_cols = {col_lc_id, col_grade, col_quarter, col_key_stage, col_domain,
                       col_subdomain, col_topic, col_lc, col_cs, col_ps, col_bloom,
                       col_type, col_tags}
        for i, val in enumerate(row):
            if i not in mapped_cols and val and i < len(headers):
                extra[headers[i]] = val

        c.execute("""
            INSERT INTO learning_competencies
            (subject_id, lc_id, grade, quarter, key_stage, domain, subdomain,
             content_topic, learning_competency, content_standard, performance_standard,
             blooms_level, competency_type, ai_tags, extra_data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            subject_id, get(col_lc_id),
            get(col_grade) or last_main_data.get("grade", ""),
            get(col_quarter) or last_main_data.get("quarter", ""),
            get(col_key_stage) or last_main_data.get("key_stage", ""),
            get(col_domain) or last_main_data.get("domain", ""),
            get(col_subdomain),
            get(col_topic) or last_main_data.get("topic", ""),
            lc_text,
            get(col_cs) or last_main_data.get("cs", ""),
            get(col_ps) or last_main_data.get("ps", ""),
            get(col_bloom), get(col_type), get(col_tags),
            json.dumps(extra, ensure_ascii=False) if extra else None
        ))
        count += 1

    return count


def load_standards(conn, subject_id, ws):
    """Parse and load standards from a worksheet."""
    rows = _read_sheet_rows(ws)
    if len(rows) < 2:
        return 0

    headers = rows[0]
    c = conn.cursor()
    count = 0
    for row in rows[1:]:
        if not any(row):
            continue
        extra = {headers[i]: row[i] for i in range(len(row)) if i < len(headers) and row[i]}
        desc = " | ".join(v for v in row if v)
        c.execute("""
            INSERT INTO standards (subject_id, standard_type, grade, key_stage, description, extra_data)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (subject_id, "content", "", "", desc, json.dumps(extra, ensure_ascii=False)))
        count += 1
    return count


def load_pedagogical_approaches(conn, subject_id, ws):
    """Parse and load pedagogical approaches from a worksheet."""
    rows = _read_sheet_rows(ws)
    if len(rows) < 2:
        return 0

    headers = rows[0]
    col_name = _find_column(headers, "approach", "name", "strategy", "method")
    col_desc = _find_column(headers, "description", "definition", "overview")

    c = conn.cursor()
    count = 0
    for row in rows[1:]:
        if not any(row):
            continue

        name = row[col_name] if col_name is not None and col_name < len(row) else row[0] if row else ""
        desc = row[col_desc] if col_desc is not None and col_desc < len(row) else ""

        if not name:
            continue

        extra = {headers[i]: row[i] for i in range(len(row)) if i < len(headers) and row[i]}
        c.execute("""
            INSERT INTO pedagogical_approaches (subject_id, approach_name, description, strategies, extra_data)
            VALUES (?, ?, ?, ?, ?)
        """, (subject_id, name, desc, "", json.dumps(extra, ensure_ascii=False)))
        count += 1
    return count


def load_21st_century_skills(conn, subject_id, ws):
    """Parse and load 21st century skills from a worksheet."""
    rows = _read_sheet_rows(ws)
    if len(rows) < 2:
        return 0

    headers = rows[0]
    col_name = _find_column(headers, "skill", "name")
    col_cat = _find_column(headers, "category", "cluster", "domain")
    col_desc = _find_column(headers, "description", "definition")

    c = conn.cursor()
    count = 0
    for row in rows[1:]:
        if not any(row):
            continue

        name = row[col_name] if col_name is not None and col_name < len(row) else row[0] if row else ""
        if not name:
            continue

        cat = row[col_cat] if col_cat is not None and col_cat < len(row) else ""
        desc = row[col_desc] if col_desc is not None and col_desc < len(row) else ""

        extra = {headers[i]: row[i] for i in range(len(row)) if i < len(headers) and row[i]}
        c.execute("""
            INSERT INTO twenty_first_century_skills (subject_id, skill_name, category, description, extra_data)
            VALUES (?, ?, ?, ?, ?)
        """, (subject_id, name, cat, desc, json.dumps(extra, ensure_ascii=False)))
        count += 1
    return count


def load_crosscutting_concepts(conn, subject_id, ws):
    """Parse and load crosscutting concepts from a worksheet."""
    rows = _read_sheet_rows(ws)
    if len(rows) < 2:
        return 0

    headers = rows[0]
    col_name = _find_column(headers, "concept", "name", "big_idea", "theme")
    col_desc = _find_column(headers, "description", "definition", "explanation")

    c = conn.cursor()
    count = 0
    for row in rows[1:]:
        if not any(row):
            continue

        name = row[col_name] if col_name is not None and col_name < len(row) else row[0] if row else ""
        if not name:
            continue

        desc = row[col_desc] if col_desc is not None and col_desc < len(row) else ""

        extra = {headers[i]: row[i] for i in range(len(row)) if i < len(headers) and row[i]}
        c.execute("""
            INSERT INTO crosscutting_concepts (subject_id, concept_name, description, extra_data)
            VALUES (?, ?, ?, ?)
        """, (subject_id, name, desc, json.dumps(extra, ensure_ascii=False)))
        count += 1
    return count


def load_domain_sequence(conn, subject_id, ws):
    """Parse and load domain sequence map from a worksheet."""
    rows = _read_sheet_rows(ws)
    if len(rows) < 2:
        return 0

    headers = rows[0]
    c = conn.cursor()
    count = 0
    for row in rows[1:]:
        if not any(row):
            continue
        extra = {headers[i]: row[i] for i in range(len(row)) if i < len(headers) and row[i]}
        domain_name = row[0] if row else ""
        c.execute("""
            INSERT INTO domain_sequence (subject_id, domain_name, grade, sequence_info, extra_data)
            VALUES (?, ?, ?, ?, ?)
        """, (subject_id, domain_name, "", " | ".join(v for v in row if v),
              json.dumps(extra, ensure_ascii=False)))
        count += 1
    return count


def load_single_subject(subject_id, filename):
    """Load all data for a single subject from its Excel file."""
    filepath = os.path.join(DATA_DIR, filename)
    if not os.path.exists(filepath):
        print(f"  WARNING: File not found: {filepath}")
        return

    print(f"  Loading {subject_id} from {filename}...")
    wb = load_workbook(filepath, read_only=True, data_only=True)

    conn = sqlite3.connect(DB_PATH)

    # Insert subject record
    conn.execute("INSERT OR REPLACE INTO subjects (id, display_name, filename) VALUES (?, ?, ?)",
                 (subject_id, SUBJECT_DISPLAY.get(subject_id, subject_id), filename))

    # Find and load each sheet type
    lc_sheet = _find_sheet(wb, ["competenc", "learning"])
    if lc_sheet:
        n = load_learning_competencies(conn, subject_id, lc_sheet)
        print(f"    Learning Competencies: {n} rows")

    std_sheet = _find_sheet(wb, ["standard"])
    if std_sheet:
        n = load_standards(conn, subject_id, std_sheet)
        print(f"    Standards: {n} rows")

    ped_sheet = _find_sheet(wb, ["pedagog", "approach"])
    if ped_sheet:
        n = load_pedagogical_approaches(conn, subject_id, ped_sheet)
        print(f"    Pedagogical Approaches: {n} rows")

    skills_sheet = _find_sheet(wb, ["21st", "century", "skill"])
    if skills_sheet:
        n = load_21st_century_skills(conn, subject_id, skills_sheet)
        print(f"    21st Century Skills: {n} rows")

    cc_sheet = _find_sheet(wb, ["crosscut", "big_idea", "concept", "theme"])
    if cc_sheet:
        n = load_crosscutting_concepts(conn, subject_id, cc_sheet)
        print(f"    Crosscutting Concepts: {n} rows")

    dom_sheet = _find_sheet(wb, ["domain", "sequence", "map"])
    if dom_sheet:
        n = load_domain_sequence(conn, subject_id, dom_sheet)
        print(f"    Domain Sequence: {n} rows")

    conn.commit()
    conn.close()
    wb.close()


def load_shs_curriculum():
    """Load all 6 SHS Core Subject competencies from the single multi-subject Excel file.

    The file has title/subtitle rows at index 0-1, real headers at index 2,
    and data starting at index 3. All subjects share one sheet, identified
    by the Subject_Code column (EC, MK, GM, GS, LCS, KLP).
    """
    filepath = os.path.join(DATA_DIR, SHS_FILE)
    if not os.path.exists(filepath):
        print(f"  WARNING: SHS file not found: {filepath}")
        return

    print(f"  Loading SHS Core Subjects from {SHS_FILE}...")
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Find the learning competencies sheet (S1_Learning_Competencies or similar)
    ws = None
    for name in wb.sheetnames:
        lower = name.lower()
        if "competenc" in lower or "learning" in lower or lower.startswith("s1"):
            ws = wb[name]
            break
    if ws is None:
        print("  WARNING: Could not find competencies sheet in SHS file.")
        wb.close()
        return

    # rows[0] = title, rows[1] = subtitle, rows[2] = header, rows[3+] = data
    rows = _read_sheet_rows(ws)
    if len(rows) < 4:
        print("  WARNING: SHS sheet has insufficient rows.")
        wb.close()
        return

    headers = rows[2]

    col_code    = _find_column(headers, "subject_code", "subject code")
    col_grade   = _find_column(headers, "grade_level", "grade level", "grade")
    col_quarter = _find_column(headers, "quarter")
    col_lc_id   = _find_column(headers, "lc_id", "lc_code")
    col_domain  = _find_column(headers, "domain", "strand", "component")
    col_lc      = _find_column(headers, "competency_statement", "learning_competency", "competency")
    col_bloom   = _find_column(headers, "bloom", "blooms")
    col_tags    = _find_column(headers, "ai_tag", "ai_searchable", "tags")

    if col_code is None:
        print("  WARNING: Could not find Subject_Code column in SHS sheet.")
        wb.close()
        return

    def normalize_grade(g):
        m = re.search(r'\d+', str(g).strip())
        return m.group() if m else str(g).strip()

    def normalize_quarter(q):
        m = re.search(r'\d+', str(q).strip())
        return m.group() if m else str(q).strip()

    # Group data rows by Subject_Code
    subject_rows = defaultdict(list)
    for row in rows[3:]:
        if not any(row):
            continue
        code = row[col_code] if col_code < len(row) else ""
        if code and code in SHS_SUBJECT_CODES:
            subject_rows[code].append(row)

    conn = sqlite3.connect(DB_PATH)
    mapped_cols = {col_code, col_grade, col_quarter, col_lc_id,
                   col_domain, col_lc, col_bloom, col_tags}

    for code, s_rows in subject_rows.items():
        subject_id   = SHS_SUBJECT_CODES[code]
        display_name = SHS_DISPLAY[subject_id]

        conn.execute(
            "INSERT OR REPLACE INTO subjects (id, display_name, filename) VALUES (?, ?, ?)",
            (subject_id, display_name, SHS_FILE)
        )

        c = conn.cursor()
        count = 0
        for row in s_rows:
            def get(idx):
                if idx is not None and idx < len(row):
                    return row[idx]
                return ""

            lc_text = get(col_lc)
            if not lc_text:
                continue

            grade   = normalize_grade(get(col_grade))   if col_grade   is not None else ""
            quarter = normalize_quarter(get(col_quarter)) if col_quarter is not None else ""

            extra = {}
            for i, val in enumerate(row):
                if i not in mapped_cols and val and i < len(headers):
                    extra[headers[i]] = val

            c.execute("""
                INSERT INTO learning_competencies
                (subject_id, lc_id, grade, quarter, key_stage, domain, subdomain,
                 content_topic, learning_competency, content_standard, performance_standard,
                 blooms_level, competency_type, ai_tags, extra_data)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                subject_id, get(col_lc_id),
                grade, quarter,
                "SHS",
                get(col_domain), "",
                "", lc_text,
                "", "",
                get(col_bloom), "",
                get(col_tags),
                json.dumps(extra, ensure_ascii=False) if extra else None
            ))
            count += 1

        print(f"    {display_name} ({code}): {count} competencies")

    conn.commit()
    conn.close()
    wb.close()


def load_all_curriculum_data():
    """Load all curriculum data from all Excel files into the database."""
    print("Initializing database...")
    init_database()

    print(f"Loading {len(SUBJECT_FILES)} MATATAG subject files...")
    for subject_id, filename in SUBJECT_FILES.items():
        load_single_subject(subject_id, filename)

    print("Loading SHS Core Subjects...")
    load_shs_curriculum()

    # Print summary
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM learning_competencies")
    total_lc = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT subject_id) FROM learning_competencies")
    total_subj = c.fetchone()[0]
    conn.close()

    print(f"\nDone! Loaded {total_lc} learning competencies across {total_subj} subjects.")
    return total_lc


# --- Query functions used by the app ---

def get_db():
    """Get a database connection."""
    if not os.path.exists(DB_PATH):
        load_all_curriculum_data()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_subjects():
    """Get all available subjects."""
    conn = get_db()
    rows = conn.execute("SELECT id, display_name FROM subjects ORDER BY display_name").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _grade_sort_key(g):
    """Sort grades numerically (K=0, 1-10)."""
    g = g.strip()
    if g.upper() in ("K", "KINDER", "KINDERGARTEN"):
        return 0
    try:
        return int(g)
    except ValueError:
        return 999


def get_grades_for_subject(subject_id):
    """Get all available grades for a given subject."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT grade FROM learning_competencies WHERE subject_id = ? AND grade != '' ORDER BY grade",
        (subject_id,)
    ).fetchall()
    conn.close()
    grades = [r["grade"] for r in rows]
    grades.sort(key=_grade_sort_key)
    return grades


def get_quarters_for_subject_grade(subject_id, grade):
    """Get all available quarters for a given subject and grade."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT quarter FROM learning_competencies WHERE subject_id = ? AND grade = ? AND quarter != '' ORDER BY quarter",
        (subject_id, grade)
    ).fetchall()
    conn.close()
    return [r["quarter"] for r in rows]


def get_competencies(subject_id, grade=None, quarter=None):
    """Get learning competencies filtered by subject, grade, and quarter."""
    conn = get_db()
    query = "SELECT * FROM learning_competencies WHERE subject_id = ?"
    params = [subject_id]

    if grade:
        query += " AND grade = ?"
        params.append(grade)
    if quarter:
        query += " AND quarter = ?"
        params.append(quarter)

    query += " ORDER BY lc_id"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_competency_by_id(comp_id):
    """Get a single learning competency by its database ID."""
    conn = get_db()
    row = conn.execute("SELECT * FROM learning_competencies WHERE id = ?", (comp_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def get_pedagogical_approaches(subject_id):
    """Get pedagogical approaches for a subject."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM pedagogical_approaches WHERE subject_id = ?", (subject_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# Fallback used when a subject's spreadsheet has no recoverable skill names (e.g. English, Filipino)
_UNIVERSAL_21C_SKILLS = [
    {"skill_name": "Critical Thinking", "category": "Learning & Innovation",
     "description": "Analyze, evaluate and synthesize information to solve complex problems."},
    {"skill_name": "Creativity", "category": "Learning & Innovation",
     "description": "Generate new ideas and approaches to learning challenges."},
    {"skill_name": "Communication", "category": "Life & Career",
     "description": "Express ideas clearly and effectively in written, oral, and multimodal forms."},
    {"skill_name": "Collaboration", "category": "Life & Career",
     "description": "Work effectively with others toward shared learning goals."},
    {"skill_name": "Information Literacy", "category": "Information, Media & Technology",
     "description": "Access, evaluate, and use information effectively and ethically."},
    {"skill_name": "Technology Literacy", "category": "Information, Media & Technology",
     "description": "Use digital tools and technology responsibly for learning and communication."},
]

# Regex for ID codes like CS01, 21C-01, 21C01, 21C1
_CODE_RE = re.compile(r'^[A-Z]{2,4}[\-_]?\d{1,3}$|^\d{2}[A-Z]{1,2}[\-_]?\d{1,3}$', re.IGNORECASE)

# Column header strings that are meta-labels, not skill names
_HEADER_LABELS = {
    'skill code', 'skill category', 'skill domain', 'skill_id',
    'skill name', 'skill sub-category', 'category', 'specific skill',
    'skill_name', 'skill_category',
}


def get_21st_century_skills(subject_id):
    """Get 21st century skills for a subject, with smart name extraction and deduplication.

    Different spreadsheets store skills differently:
    - Some use codes (CS01, 21C-01) as skill_name → real name is in extra_data
    - Some store the category as skill_name and repeat it for each sub-skill
    - Some have a 'Specific Skill' or 'Skill Sub-Category' field in extra_data
    """
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM twenty_first_century_skills WHERE subject_id = ?", (subject_id,)
    ).fetchall()
    conn.close()

    skills = []
    seen_names = set()

    for row in rows:
        r = dict(row)
        raw_name = (r.get('skill_name') or '').strip()
        extra = json.loads(r.get('extra_data') or '{}')

        # Skip header rows that leaked into the data
        if raw_name.lower() in _HEADER_LABELS:
            continue

        name = raw_name

        # Priority 1: Specific Skill (AP, Reading_Literacy, Mathematics, PE_Health)
        specific = extra.get('Specific Skill', '').strip()
        if specific and not _CODE_RE.match(specific) and specific.lower() not in _HEADER_LABELS:
            name = specific

        # Priority 2: Skill Sub-Category (Makabansa)
        elif not specific:
            sub = extra.get('Skill Sub-Category', '').strip()
            if sub and not _CODE_RE.match(sub) and sub.lower() not in _HEADER_LABELS:
                name = sub

        # Priority 3: Skill Name / Skill_Name (Kindergarten, Music_Arts — code stored in skill_name)
        if _CODE_RE.match(name):
            for key in ('Skill Name', 'Skill_Name'):
                val = extra.get(key, '').strip()
                if val and not _CODE_RE.match(val) and val.lower() not in _HEADER_LABELS:
                    name = val
                    break

        # Still a code or empty → skip (e.g. English CS01, Filipino 21C-01 with no recovery)
        if not name or _CODE_RE.match(name):
            continue

        # Deduplicate case-insensitively
        name_key = name.lower()
        if name_key in seen_names:
            continue
        seen_names.add(name_key)

        # Get description — prefer stored description, fall back to extra_data
        desc = (r.get('description') or '').strip()
        if not desc:
            desc = extra.get('Description', extra.get('description', '')).strip()

        # Get category
        cat = (r.get('category') or '').strip()
        if not cat:
            cat = extra.get('Skill Category', extra.get('Skill_Category', raw_name)).strip()

        skills.append({
            'id': r['id'],
            'subject_id': r['subject_id'],
            'skill_name': name,
            'category': cat,
            'description': desc,
        })

    # Fallback for subjects whose spreadsheet has no recoverable skill names (e.g. English, Filipino)
    if not skills:
        return [dict(s, id=0, subject_id=subject_id) for s in _UNIVERSAL_21C_SKILLS]

    return skills


def get_crosscutting_concepts(subject_id):
    """Get crosscutting concepts for a subject."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM crosscutting_concepts WHERE subject_id = ?", (subject_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


if __name__ == "__main__":
    load_all_curriculum_data()
